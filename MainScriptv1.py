#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#importing packages
flag=0
import requests
import os

from datetime import datetime
import time

#Download dose json file
file_url=file_url=os.environ.get("URL1")
  
r = requests.get(file_url, stream = True)
  
with open("Dose.json","wb") as pdf:
     flag+=1
     for chunk in r.iter_content(chunk_size=1024):
          # writing one chunk at a time to pdf file
          if chunk:
               pdf.write(chunk)


#download recover json file

file_url=os.environ.get("URL2")
  
r = requests.get(file_url, stream = True)
  
with open("Recover2.json","wb") as pdf:
    flag+=1
    for chunk in r.iter_content(chunk_size=1024):
  
         # writing one chunk at a time to pdf file
         if chunk:
             pdf.write(chunk)



#convert json to csv Dose file


#----------------------------------------------------
import json
import pandas as pd
import csv
import numpy as np

with open("Dose.json") as file:
    data = json.load(file)

fname = "Dose.csv"

with open(fname, "w") as file:
    flag+=1
    csv_file = csv.writer(file,lineterminator='\n')
    csv_file.writerow(["Date","State","Dose 1","Dose 2","Total Doses"])
    for item in data["vaccine_data"]:
        for item2 in item["vacc_st_data"]:
            csv_file.writerow([item['day'],item2['st_name'],item2['dose1'],
                               item2['dose2'],item2['total_doses']])


# making data frame from csv file
data = pd.read_csv("Dose.csv")
  
# sorting data frame by multiple columns
data.sort_values(["State","Date"], axis=0,
                 ascending=True, inplace=True)
  
data.to_csv('Dose.csv', index=False)

#-------------------------------------




#remove the hindi characters from confirmed json as they are not supported by openpyxl

#--------------------------------------------------------
import json

with open('Recover2.json', 'r') as jf:
    jsonFile = json.load(jf)

testJson = {}
keyList = jsonFile.keys()
count=0
for key in keyList:
    if count==3:
        count+=1
        continue
    else:
        testJson[key] = jsonFile[key]
        count+=1
        
with open('Recover.json', 'w') as jf:
    flag+=1
    json.dump(testJson, jf)


#-----------------------------------------------------------


#convert json to csv Recover file

#-----------------------------------------------------------

import pandas as pd

# load json file using pandas
df = pd.read_json('Recover.json')


# convert dataframe to csv file
df.to_csv("Recover.csv",index=False)
flag+=1

#---------------------------------------------------------


#convert csv to xlsx dose file (change name) 

#---------------------------------------------------------
import json
import pandas as pd
import csv
import numpy as np  
  
# Reading the csv file
df_new = pd.read_csv('Dose.csv')
  
# saving xlsx file
GFG = pd.ExcelWriter('Dose.xlsx')
df_new.to_excel(GFG, index = False)
  
GFG.save()
flag+=1
#-----------------------------------------------------------



#convert csv to xlsx recoverd file
#--------------------------------------------------------------
import json
import pandas as pd
import csv
import numpy as np
# Reading the csv file
df_new = pd.read_csv('Recover.csv')
  
# saving xlsx file
GFG = pd.ExcelWriter('Recover.xlsx')
df_new.to_excel(GFG, index = False)
  
GFG.save()
flag+=1
#-------------------------------------------------------------
import os
from openpyxl import*
from openpyxl.utils import*
import sys
from datetime import datetime
import time
#Main writer program

start = time.time()

current_time = datetime.now()

DateForMaster=("%s/%s/%s" % (current_time.month, current_time.day, current_time.year))#3/31/2020
    
    
Day=current_time.day
Month=current_time.month
Year=current_time.year
if (len(str(current_time.month))==1 and len(str(current_time.day))==1):
    DateForDose=("%s-0%s-0%s" % (current_time.year,current_time.month,current_time.day-1))#2021-03-03
        
elif(len(str(current_time.day))==1):
    DateForDose=("%s-%s-0%s" % (current_time.year,current_time.month,current_time.day-1))#2021-03-03
        
elif(len(str(current_time.month))==1):
    DateForDose=("%s-0%s-%s" % (current_time.year,current_time.month,current_time.day-1))#2021-03-03
        
else:
    DateForDose=("%s-%s-%s" % (current_time.year,current_time.month,current_time.day-1))#2021-03-03






wb = load_workbook('Master Sheet.xlsx')
ws = wb.active

wbDose = load_workbook('Dose.xlsx')
wsDose = wbDose.active

wbRecover = load_workbook('Recover.xlsx')
wsRecover = wbRecover.active

#check condition for data is already upadted or not

for xyz in range(2,ws.max_row):
    if (ws['A'+str(xyz)].value==DateForMaster):
        print ("records already updated ")
        
        os.remove("Dose.csv")
        os.remove("Dose.json")
        os.remove("Dose.xlsx")
        os.remove("Recover.csv")
        os.remove("Recover.json")
        os.remove("Recover.xlsx")
        os.remove("Recover2.json")

        sys.exit()
        break




#checking the data is available or not
#for dose
for d in range(1,wsDose.max_row):
    if (DateForDose==wsDose['A'+str(d)].value):
        flag+=1
        break



#just for reference
'''list_all=["Andaman and Nicobar Islands","Andhra Pradesh","Arunachal Pradesh","Assam","Bihar","Chandigarh",
       "Chhattisgarh","Dadra and Nagar Haveli and Daman and Diu","Delhi","Goa","Gujarat","Haryana",
       "Himachal Pradesh","India","Jammu and Kashmir","Jharkhand","Karnataka","Kerala","Ladakh",
       "Lakshadweep","Madhya Pradesh","Maharashtra","Manipur","Meghalaya","Mizoram","Nagaland","Odisha",
       "Puducherry","Punjab","Rajasthan","Sikkim","State Unassigned","Tamil Nadu",
       "Telangana","Tripura","Uttar Pradesh","Uttarakhand",
       "West Bengal","Miscellaneous"]'''



State_list=["Andaman and Nicobar Islands","Andhra Pradesh","Arunachal Pradesh","Assam","Bihar","Chandigarh",
       "Chhattisgarh","Dadra and Nagar Haveli and Daman and Diu","Delhi","Goa","Gujarat","Haryana",
       "Himachal Pradesh","India","Jammu and Kashmir","Jharkhand","Karnataka","Kerala","Ladakh",
       "Lakshadweep","Madhya Pradesh","Maharashtra","Manipur","Meghalaya","Mizoram","Nagaland","Odisha",
       "Puducherry","Punjab","Rajasthan","Sikkim","Tamil Nadu",
       "Telangana","Tripura","Uttar Pradesh","Uttarakhand",
       "West Bengal"]

#Renaming Telengana to Telangana in Recover file
for j in range(1,wsRecover.max_row):
    if (wsRecover['A'+str(j)].value=="Telengana"):
        wsRecover['A'+str(j)].value="Telangana"
wbRecover.save('Recover.xlsx')

#renaming Andaman and Nicobar to Andaman and Nicobar Islands in Dose file
for j in range(1,wsDose.max_row):
    if (wsDose['B'+str(j)].value=="Andaman and Nicobar"):
        wsDose['B'+str(j)].value="Andaman and Nicobar Islands"
wbDose.save('Dose.xlsx')


# insert rows for data insertion
x=0
for i in range(2,ws.max_row):
    data=(ws['B'+str(i)].value)
    if (data!=State_list[x] and (data in State_list)):
        ws.insert_rows(i)
        ws['A'+str(i)].value=DateForMaster
        #print ("Insert cell Successfull before ",data)
        
        if (x<36):
            x+=1
        else:
            break


#del row
for x in range(2,ws.max_row):
    if (ws['B'+str(x)].value=="Tamil Nadu"):
        ws.delete_rows(x-1)
        break

for x in range(2,ws.max_row):
    if (ws['B'+str(x)].value=="State Unassigned"):
        ws.insert_rows(x)
        ws['A'+str(x)].value=DateForMaster
        break


    


Start=1
for i in range(2,ws.max_row):
    if ws['B'+str(i)].value==None:
        data=ws['B'+str(i-5)].value
        ws['B'+str(i)].value=data
        
        for j in range(1,wsRecover.max_row):
            if (data==(wsRecover['A'+str(j)].value)):
                ws['C'+str(i)].value=wsRecover['H'+str(j)].value
                ws['D'+str(i)].value=wsRecover['J'+str(j)].value
                ws['E'+str(i)].value=wsRecover['K'+str(j)].value
                ws['F'+str(i)].value=0
                #print ("Data1 inserted for ",data)
                
                

        for y in range(Start,wsDose.max_row):
            if (wsDose['B'+str(y)].value==data and DateForDose==wsDose['A'+str(y)].value):
                ws['I'+str(i)].value=wsDose['C'+str(y)].value
                ws['J'+str(i)].value=wsDose['D'+str(y)].value
                ws['H'+str(i)].value=wsDose['E'+str(y)].value
                #print ("Data2222 inserted for ",data)
                Start=y-5
                        




#westbangal left

lastrow=ws.max_row

ws['B'+str(lastrow+1)].value=ws['B'+str(ws.max_row-5)].value
ws['A'+str(lastrow+1)].value=DateForMaster
for j in range(2,wsRecover.max_row+8):
    if (wsRecover['A'+str(j)].value=="West Bengal"):
        
        ws['C'+str(lastrow+1)].value=wsRecover['H'+str(j)].value
        ws['D'+str(lastrow+1)].value=wsRecover['J'+str(j)].value
        ws['E'+str(lastrow+1)].value=wsRecover['K'+str(j)].value
        ws['F'+str(lastrow+1)].value=0
        #print ("Data1 inserted for westbangal")
                
                

for y in range(2,wsDose.max_row+8):
    if (wsDose['B'+str(y)].value=="West Bengal" and wsDose['A'+str(y)].value==DateForDose):
        
        ws['I'+str(lastrow+1)].value=wsDose['C'+str(y)].value
        ws['J'+str(lastrow+1)].value=wsDose['D'+str(y)].value
        ws['H'+str(lastrow+1)].value=wsDose['E'+str(y)].value
        #print ("Data2222 inserted for west bangal")
#for andaman

for x in range(2,ws.max_row):
     if (ws['A'+str(x)].value==DateForMaster):
          trace=x
          #print("found ")
          break
for xt in range(2,wsRecover.max_row+8):
    if (wsRecover['A'+str(xt)].value=="Andaman and Nicobar"):
        
        ws['C'+str(x)].value=wsRecover['H'+str(xt)].value
        ws['D'+str(x)].value=wsRecover['J'+str(xt)].value
        ws['E'+str(x)].value=wsRecover['K'+str(xt)].value
        ws['F'+str(x)].value=0
        #print ("Data1 inserted for Andaman and Nicobar Islands")




wb.save('Master Sheet.xlsx')


#del india extra entry
for xy in range(2,ws.max_row):
    if ((ws['B'+str(xy)].value=="India") and (ws['A'+str(xy)].value==DateForMaster)):
        ws.delete_rows(xy)
        break



wb.save('Master Sheet.xlsx')    
print("Time taken by Program")
elapsed_time_fl = (time.time() - start)
print(elapsed_time_fl)

#deleting temp files

import os
os.remove("Dose.csv")
os.remove("Dose.json")
os.remove("Dose.xlsx")
os.remove("Recover.csv")
os.remove("Recover.json")
os.remove("Recover.xlsx")
os.remove("Recover2.json")

#data inconsistency checks


for x in range(1,ws.max_row):
    if (ws['A'+str(x)].value==DateForMaster):
        if ((ws['C'+str(x)].value=="") or (ws['D'+str(x)].value=="") or (ws['E'+str(x)].value=="") or 
            (ws['H'+str(x)].value=="") or (ws['I'+str(x)].value=="") or (ws['J'+str(x)].value=="")):
            print("Data Inconsistency found .........................................!!!!!!!!!!!!!!!")
            break
